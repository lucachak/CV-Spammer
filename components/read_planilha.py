from openpyxl import load_workbook

DATA = load_workbook("aplicacoes_catho.xlsx").active


class Reader:

    __slots__ = ("dados", "headers")

    def __init__(self) -> None:
        self.dados = []
        self.headers = [cell.value for cell in DATA[1]]

    def read_planilha(self) -> None:
        for row in DATA.iter_rows(min_row=2, values_only=True):
            self.dados.append(dict(zip(self.headers, row)))
        [print(x) for x in self.dados]
